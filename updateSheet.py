# import sqlite3, requests, json
# import pandas as pd
# from openpyxl import load_workbook



# # url="https://script.google.com/macros/s/AKfycbwXuYUci3bglvDKQ-KS3yczqKXzO_ky5bbqskr4O_H8prZgqmdnzxWv8KwjK2SCSgTx/exec"

# def get_data(min, max):
#     # return the rows whose matchNum is within the range [min, max] inclusive

#     conn = sqlite3.connect('data/scouting_dat.db')
#     c = conn.cursor()
#     c.execute('SELECT * FROM scoutData WHERE matchNum BETWEEN ? AND ?;', (min, max))
#     rows = c.fetchall() 
#     print(rows)
#     conn.close()

#     # Format data
#     to_send = []
#     scout_rows = []  # Keep track of original scout rows
#     for row in rows:
#         print(row, "row beta")
#         to_send.append(json.loads(row[4])) # Extract information in the 'data' column
#         to_send[-1]["timestamp"] = row[0] # Append metadata
#         to_send[-1]["matchNum"] = row[1]
#         to_send[-1]["teamNum"] = row[2]
#         to_send[-1]["scoutID"] = row[3]
#         to_send[-1]["type"] = "scout"
#         scout_rows.append(row)  # Store original row
#     print(to_send)
#     return to_send, scout_rows

# def move_to_old(scout_rows):
#     if not scout_rows:
#         return  # No scout data to move

#     # Connect to old database
#     conn_old = sqlite3.connect('data/scouting_dat_old.db')
#     c_old = conn_old.cursor()
    
#     # Create table if not exists (assuming same schema as scoutData)
#     c_old.execute('''CREATE TABLE IF NOT EXISTS scoutData (
#         timestamp TEXT,
#         matchNum INTEGER,
#         teamNum INTEGER,
#         scoutID TEXT,
#         data TEXT
#     )''')
    
#     # Insert rows into old database
#     c_old.executemany('INSERT INTO scoutData VALUES (?, ?, ?, ?, ?)', scout_rows)
#     conn_old.commit()
#     conn_old.close()
    
#     # Delete from original database
#     conn = sqlite3.connect('data/scouting_dat.db')
#     c = conn.cursor()
#     # Delete the specific rows that were sent
#     for row in scout_rows:
#         c.execute('DELETE FROM scoutData WHERE timestamp = ? AND matchNum = ? AND teamNum = ? AND scoutID = ? AND data = ?',
#                   (row[0], row[1], row[2], row[3], row[4]))
#     conn.commit()
#     conn.close()

# def send_match(matchNum):
#     # send_match: fetches data from one match and 
#     print(f"Uploading data from match {matchNum}... ", end="")

#     data, scout_rows = get_data(matchNum, matchNum)
#     print("----DATA STARTS HERE-----")
#     print(data)
#     print("----DATA ENDS HERE-----" )

#     if len(data) == 0:
#         print("Failed")
#         return f"Error: No data found for match {matchNum}"

#     for i in range(len(data)):
#         # print(i)

#         wb = load_workbook('scoutingDataRAW.xlsx')
#         ws = wb.active
#         # try:
#         print(data[i])
#         # print(data[i]["failureCheck"])

#         if (data[i]["failureCheck"] == "y"):
#             data[i]["failureCheck"] = "y"
#         else:
#             data[i]["failureCheck"] = "n"
#             data[i]["failureCheck"] = " "


#         ws.append([
#             data[i]["timestamp"],
#             data[i]["scoutID"],
#             data[i]["matchNum"],
#             data[i]["teamNum"],
#             data[i]["autoShots"],
#             data[i]["totalPositions"],
#             data[i]["heatmap_frequencies"],
#             data[i]["teleopShots"],
#             data[i]["passes"],
#             data[i]["climbLocation"],
#             data[i]["fouls"],
#             data[i]["failureDetails"],
#             data[i]["failureCheck"],
#             data[i]["climbTime"],
#             data[i]["info"]
#         ])
            
#         # except:
#         #     print("Failed")
#         #     return "Error: POST request failed"
#     wb.save('scoutingDataRAW.xlsx')

#     # # status = response.json()
#     # if status['status'] != 'success':
#     #     print(f"App script returned failing status: {status}")
#     #     return "Error: App script returned failing status"
    

#     print("Success")
#     # Move data from scouting_dat to scouting_dat_old after successful send

#     # move_to_old(scout_rows)
#     return f"Sent"


import sqlite3
import json
from openpyxl import load_workbook
import os


def get_data(min_match, max_match):
    conn = sqlite3.connect('data/scouting_dat.db')
    c = conn.cursor()

    c.execute(
        'SELECT * FROM scoutData WHERE matchNum BETWEEN ? AND ?;',
        (min_match, max_match)
    )
    rows = c.fetchall()
    conn.close()

    to_send = []
    scout_rows = []

    for row in rows:
        try:
            data = json.loads(row[4])
        except:
            continue  # skip bad rows

        data["timestamp"] = row[0]
        data["matchNum"] = row[1]
        data["teamNum"] = row[2]
        data["scoutID"] = row[3]
        data["type"] = "scout"

        to_send.append(data)
        scout_rows.append(row)

    return to_send, scout_rows


def move_to_old(scout_rows):
    if not scout_rows:
        return

    conn_old = sqlite3.connect('data/scouting_dat_old.db')
    c_old = conn_old.cursor()

    c_old.execute('''CREATE TABLE IF NOT EXISTS scoutData (
        timestamp TEXT,
        matchNum INTEGER,
        teamNum INTEGER,
        scoutID TEXT,
        data TEXT
    )''')

    c_old.executemany(
        'INSERT INTO scoutData VALUES (?, ?, ?, ?, ?)',
        scout_rows
    )

    conn_old.commit()
    conn_old.close()

    conn = sqlite3.connect('data/scouting_dat.db')
    c = conn.cursor()

    for row in scout_rows:
        c.execute('''
            DELETE FROM scoutData
            WHERE timestamp=? AND matchNum=? AND teamNum=? AND scoutID=? AND data=?
        ''', row)

    conn.commit()
    conn.close()


def send_match(matchNum):
    print(f"Uploading data from match {matchNum}...")

    data, scout_rows = get_data(matchNum, matchNum)

    print("DATA FOUND:", len(data))

    if len(data) == 0:
        print("No data found.")
        return

    # Check file exists
    file_path = 'ManorOfficialScouting2026.xlsx'

    if not os.path.exists(file_path):
        print("Excel file not found!")
        return

    # LOAD WORKBOOK ONCE
    wb = load_workbook(file_path)
    ws = wb["raw match data"]

    for entry in data:
        try:
            failure = "y" if entry.get("failureCheck") == "y" else ""
            autonClimb = "y" if entry.get("autonClimb") == "y" else ""
            teleopDefense = "true" if entry.get("teleopDefense") == "true" else ""
            climbFailed = "true" if entry.get("climbFailed") == "true" else ""


            ws.append([
                entry.get("timestamp", ""),
                entry.get("matchNum", ""),
                entry.get("teamNum", ""),
                entry.get("scoutID", ""), # 67
                autonClimb,
                entry.get("totalPositions", ""),
                teleopDefense,
                entry.get("successfulStops", ""),
                entry.get("climb", ""),
                climbFailed,
                entry.get("failureDetails", ""),
                failure,
                entry.get("fouls", ""),
                entry.get("info", "")
            ])

        except Exception as e:
            print("Error writing row:", e)

    wb.save(file_path)
    print("Excel updated successfully.")

    # OPTIONAL: move data after success
    # move_to_old(scout_rows)

    return "Sent"